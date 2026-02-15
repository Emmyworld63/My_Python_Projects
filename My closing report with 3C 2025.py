# build_3chub_bi_2025.py
# Rebuilds the '3chub_Retail_BA_Report_2025_FULL_with_Raw_and_Automations.xlsx' workbook from a JL export.

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference, BarChart

SRC_PATH = r"Smart Phone MSL 2025 - Copy.xlsx"
OUT_PATH = r"3chub_Retail_BA_Report_2025_FULL_with_Raw_and_Automations.xlsx"

def add_df(ws, df, start_row=1, start_col=1, table_name=None, money_cols=None, pct_cols=None, freeze_panes=None):
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for c in range(start_col, start_col + df.shape[1]):
        cell = ws.cell(row=start_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    for i, col in enumerate(df.columns, start=start_col):
        max_len = max([len(str(col))] + [len(str(v)) for v in df.iloc[:200, i-start_col].astype(str).values])
        ws.column_dimensions[get_column_letter(i)].width = min(max(10, max_len + 2), 45)

    money_cols = set(money_cols or [])
    pct_cols = set(pct_cols or [])
    for j, col in enumerate(df.columns, start=start_col):
        fmt = None
        if col in money_cols:
            fmt = '₦#,##0'
        elif col in pct_cols:
            fmt = '0.00%'
        if fmt:
            for r in range(start_row+1, start_row+1+len(df)):
                ws.cell(row=r, column=j).number_format = fmt

    if table_name:
        end_row = start_row + len(df)
        end_col = start_col + df.shape[1] - 1
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        ws.add_table(tab)

def main():
    raw = pd.read_excel(SRC_PATH)

    # Parse dates
    date_cols = ['Creating Date','Receive Date','Accounting Date','Shipping Date','Audit Date']
    for c in date_cols:
        if c in raw.columns:
            raw[c] = pd.to_datetime(raw[c], errors='coerce')

    # Remove totals/footer rows
    totals_mask = raw['Receipt No.'].isna() & raw['Product Description'].isna()
    totals_mask |= raw['Receipt No.'].astype(str).str.strip().isin(['nan','None','']) & raw['Product Description'].isna()
    clean = raw.loc[~totals_mask].copy()

    # Numeric coercion
    num_cols = ['Sale Qty','Actual Amount','Retail Unit Price','Receipt Limit Price','Original Sale Price',
                'Sale Amount Without Tax','Original Sale Price.1']
    for c in num_cols:
        if c in clean.columns:
            clean[c] = pd.to_numeric(clean[c], errors='coerce')

    # Features
    clean['Txn_Type'] = np.where(clean['Sale Type'].astype(str).str.contains('return', case=False, na=False) |
                                 (clean['Sale Qty']<0) | (clean['Actual Amount']<0), 'Return','Sale')
    clean['Txn_Date'] = clean['Accounting Date'].fillna(clean['Receive Date']).fillna(clean['Creating Date'])
    clean['Year'] = clean['Txn_Date'].dt.year
    clean['Month'] = clean['Txn_Date'].dt.to_period('M').dt.to_timestamp()
    clean['MonthKey'] = clean['Month'].dt.strftime('%Y-%m')
    clean['Quarter'] = clean['Txn_Date'].dt.to_period('Q').astype(str)

    clean['Net_Sales'] = clean['Actual Amount'].fillna(0.0)
    clean['Net_Units'] = clean['Sale Qty'].fillna(0.0)

    # Discounts (exclude placeholder list prices by cap)
    orig = clean['Original Sale Price'].where(clean['Original Sale Price']>0)
    valid_orig = orig.where(orig<=500000)
    clean['Discount_Value'] = (valid_orig - clean['Actual Amount']).where(valid_orig.notna() & clean['Actual Amount'].notna())
    clean['Discount_Pct'] = (clean['Discount_Value'] / valid_orig).where(valid_orig.notna() & valid_orig!=0)
    bins = [-np.inf, 0, 0.05, 0.10, 0.20, np.inf]
    labels = ['<=0% (No/Neg Disc)','0-5%','5-10%','10-20%','20%+']
    clean['Discount_Band'] = pd.cut(clean['Discount_Pct'], bins=bins, labels=labels)

    clean_2025 = clean.loc[clean['Year']==2025].copy()

    gross_sales = clean_2025.loc[clean_2025['Net_Sales']>0,'Net_Sales'].sum()
    return_value = -clean_2025.loc[clean_2025['Net_Sales']<0,'Net_Sales'].sum()
    net_sales = clean_2025['Net_Sales'].sum()
    gross_units = clean_2025.loc[clean_2025['Net_Units']>0,'Net_Units'].sum()
    return_units = -clean_2025.loc[clean_2025['Net_Units']<0,'Net_Units'].sum()
    net_units = clean_2025['Net_Units'].sum()
    receipts = clean_2025['Receipt No.'].nunique()
    asp = net_sales/net_units if net_units else np.nan

    kpi = pd.DataFrame([{
        'Gross Sales': gross_sales,
        'Returns (Value)': -return_value,
        'Net Sales': net_sales,
        'Gross Units': gross_units,
        'Returned Units': -return_units,
        'Net Units': net_units,
        'Transactions (Receipts)': receipts,
        'ASP (Net)': asp,
        'Return Rate (Value)': (return_value/gross_sales) if gross_sales else np.nan,
        'Return Rate (Units)': (return_units/gross_units) if gross_units else np.nan
    }])

    wb = Workbook()
    wb.remove(wb.active)
    ws_raw = wb.create_sheet("Raw_Data")
    add_df(ws_raw, raw, table_name="RawData", freeze_panes="A2")

    ws_clean = wb.create_sheet("Cleaned_2025")
    add_df(ws_clean, clean_2025, table_name="Cleaned2025", freeze_panes="A2")

    ws_kpi = wb.create_sheet("KPI")
    add_df(ws_kpi, kpi, table_name="KPI", freeze_panes="A2")

    wb.save(OUT_PATH)

if __name__ == "__main__":
    main()
